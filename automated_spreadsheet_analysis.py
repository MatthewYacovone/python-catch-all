from fileinput import filename

import openpyxl as xl
from openpyxl.chart import (BarChart, Reference)

def process_workbook(filename):
    wb = xl.load_workbook(filename) # create workbook
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1): # from the 2nd row until the last row
        cell = sheet.cell(row, 3) # grab each cell value
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4) # assign a var to a blank cell
        corrected_price_cell.value = corrected_price # fill cell

    # create a new instance of a Reference
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    # create a new instance of a chart
    chart = BarChart()
    chart.add_data(values) # use object method
    sheet.add_chart(chart, 'e2') #() # use object method

    wb.save(filename)






